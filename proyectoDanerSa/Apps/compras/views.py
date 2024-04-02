from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .models import Compra
from Apps.inventario.models import Producto
from .forms import CompraForm
from datetime import date
from decimal import Decimal, ROUND_HALF_EVEN, ROUND_HALF_UP
from django.contrib import messages





from django.utils import timezone 
from openpyxl.styles import  PatternFill, Border, Side
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

#Para PDF
from openpyxl.styles import PatternFill, Border, Side, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER




@login_required(login_url="/accounts/login/")
def sales_list_view(request):
    compras = Compra.objects.all()
    return render(request, "compras/compras.html", {'compras': compras})

@login_required(login_url="/accounts/login/")
def show_realizar_compra(request):
    compras = Compra.objects.all()
    productos = Producto.objects.all()
    inventario_info = [{'nombre': producto.nombre, 'existencia': producto.existencia} for producto in productos]
    return render(request, 'compras/realizar_compra.html', {'compras': compras, 'inventario_info': inventario_info})

@login_required(login_url="/accounts/login/")
def realizar_compra(request):
    if request.method == 'POST':
        form = CompraForm(request.POST)
        if form.is_valid():
            compra = form.save(commit=False)
            
            # Captura la fecha actual
            compra.fecha = date.today()

            # Lógica para actualizar existencia y costo promedio
            actualizar_existencia(compra)
            actualizar_costo_promedio(compra)

            compra.save()

            # Obtener información actualizada del inventario
            productos = Producto.objects.all()
            inventario_info = [{'nombre': producto.nombre, 'existencia': producto.existencia} for producto in productos]

            messages.success(request, 'Compra realizada exitosamente.',  extra_tags='success important')


            # Redirigir a la página de compras
            return redirect('Apps.compras:compras')

    else:
        form = CompraForm()

    # Obtén la fecha actual en formato "AAAA-MM-DD"
    fecha_actual = date.today().strftime('%Y-%m-%d')

    # Obtener información inicial del inventario
    productos = Producto.objects.all()
    inventario_info = [{'nombre': producto.nombre, 'existencia': producto.existencia} for producto in productos]

    return render(request, 'compras/realizar_compra.html', {'form': form, 'fecha_actual': fecha_actual, 'inventario_info': inventario_info})

def actualizar_existencia(compra):
    # Convertir cantidad a Decimal
    cantidad_decimal = Decimal(str(compra.cantidad))

    # Sumar la cantidad convertida a la existencia del producto
    compra.producto.existencia += cantidad_decimal

def actualizar_costo_promedio(compra):
    try:
        producto = Producto.objects.get(idproducto=compra.producto.idproducto)
    except Producto.DoesNotExist:
        # Manejar el caso en el que el producto no existe
        return

    # Guarda el producto con el nuevo precio_compra
    producto.save()

def agregar_producto_desde_modal(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        existencia = request.POST.get('existencia')
        precio_venta = request.POST.get('precio_venta')
        precio_compra = request.POST.get('precio_compra')
        estado = request.POST.get('estado')

        if float(precio_compra) >= float(precio_venta):
            error_message = '¡El precio de compra debe ser menor que el precio de venta!'
            messages.error(request, error_message, extra_tags='danger')  # Cambia el tag a 'danger' para mostrar el mensaje en rojo
            # Redirigir de vuelta a la vista de realizar compra
            return redirect('Apps.compras:realizar_compra_submit')
        elif Producto.objects.filter(nombre=nombre).exists():
            error_message = '¡El producto ya existe!'
            messages.error(request, error_message, extra_tags='danger')  # Cambia el tag a 'danger' para mostrar el mensaje en rojo
            # Redirigir de vuelta a la vista de realizar compra
            return redirect('Apps.compras:realizar_compra_submit')
        else:
            try:
                producto = Producto.objects.create(
                    nombre=nombre,
                    existencia=existencia,
                    precio_venta=precio_venta,
                    precio_compra=precio_compra,
                    estado=estado
                )
                messages.success(request, 'Producto agregado exitosamente.', extra_tags='success')  # Cambia el tag a 'success' para mostrar el mensaje en verde
                # Redirigir solo después de agregar el producto exitosamente
                return redirect('Apps.compras:realizar_compra_submit')
            except Exception as e:
                error_message = 'Error al crear el producto: ' + str(e)
                messages.error(request, error_message, extra_tags='danger')  # Cambia el tag a 'danger' para mostrar el mensaje en rojo
                # Redirigir de vuelta a la vista de realizar compra
                return redirect('Apps.compras:realizar_compra_submit')

    # Si la solicitud no es POST, redirigir a la página de realizar compra
    return redirect('Apps.compras:realizar_compra_submit')

@login_required(login_url="/accounts/login/")
def ComprasView(request):

    return render(request, "compras/Creport.html")    

def generar_reporte_compras_excel(request, start_date, end_date):
    start_date = timezone.datetime.strptime(start_date, '%Y-%m-%d')
    end_date = timezone.datetime.strptime(end_date, '%Y-%m-%d')

    compras = Compra.objects.filter(fecha__range=[start_date, end_date])

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=reporte_compras.xlsx'

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Reporte de Compras"

    headers = ['Fecha', 'Producto', 'Cantidad', 'Costo de Compra', 'Observaciones']

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = worksheet.cell(row=2, column=col_num, value=header)

        cell.fill = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid')
        cell.border = Border(left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'),
                            top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'))
        cell.alignment = Alignment(horizontal='center')

    for row_num, compra in enumerate(compras, 3):
        worksheet.cell(row=row_num, column=1, value=compra.fecha).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=2, value=compra.producto.nombre).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=3, value=compra.cantidad).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=4, value=compra.costo_compra).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=5, value=compra.observaciones).alignment = Alignment(horizontal='center')

    for col_num in range(1, worksheet.max_column + 1):
        max_length = 0
        column = get_column_letter(col_num)
        for cell in worksheet[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width

    workbook.save(response)

    return response

def generar_reporte_compras_pdf(request, start_date, end_date):
    start_date = timezone.datetime.strptime(start_date, '%Y-%m-%d')
    end_date = timezone.datetime.strptime(end_date, '%Y-%m-%d')

    compras = Compra.objects.filter(fecha__range=[start_date, end_date])

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=reporte_compras.pdf'

    # Create a PDF document
    pdf = SimpleDocTemplate(response, pagesize=letter)
    styles = getSampleStyleSheet()
    style_title = styles['Title']
    style_table_header = styles['Heading4']
    style_table_cell = ParagraphStyle(
        'TableCellStyle', parent=styles['Normal'], alignment=TA_CENTER)

    # Data for the table
    data = []
    data.append(['Fecha', 'Producto', 'Cantidad', 'Costo de Compra', 'Observaciones'])

    for compra in compras:
        data.append([
            compra.fecha.strftime('%Y-%m-%d'),
            compra.producto.nombre,
            str(compra.cantidad),
            str(compra.costo_compra),
            compra.observaciones
        ])

    # Create the table and style
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.gray),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
    ]))

    # Build PDF document
    elements = []
    elements.append(table)
    pdf.build(elements)

    return response


from django.http import JsonResponse
@login_required(login_url="/accounts/login/")
def obtener_datos_filtrados_compras(request, start_date, end_date):
    start_date = timezone.datetime.strptime(start_date, '%Y-%m-%d')
    end_date = timezone.datetime.strptime(end_date, '%Y-%m-%d')

    compras = Compra.objects.filter(fecha__range=[start_date, end_date])

    data = []

    for compra in compras:
        fecha = compra.fecha.strftime('%Y-%m-%d')
        producto = compra.producto.nombre
        cantidad = compra.cantidad
        costo_compra = compra.costo_compra
        observaciones = compra.observaciones

        data.append({
            'fecha': fecha,
            'producto': producto,
            'cantidad': cantidad,
            'costo_compra': costo_compra,
            'observaciones': observaciones,
        })

    return JsonResponse(data, safe=False)