from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .models import Cliente, SaldoInicial, AbonoSaldoInicial
from Apps.ventas.models import Venta, Transaccion
from Apps.rutas.models import Ruta
from django.contrib import messages
from django.db.models import Sum
from django.http import JsonResponse
from django.views import View
from django.db import transaction
from django.db.models import Q
from decimal import Decimal
from django.utils import timezone




from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from django.http import HttpResponse


from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER





@login_required(login_url="/accounts/login/")
def ListaClientesView(request):
    context = {
        "clientes": Cliente.objects.filter(activo=True).exclude(id=1),
        "rutas": Ruta.objects.filter(activo=True),
    }
    return render(request, "clientes/clientes.html", context=context)



        
@login_required(login_url="/accounts/login/")
def DetalleClientesView(request, id_cliente):
    cliente = Cliente.objects.get(pk=id_cliente)

    ventas_no_pagadas = Venta.objects.filter(id_clientes=id_cliente)
    transaccion = Transaccion.objects.all()
    abonos = AbonoSaldoInicial.objects.filter(id_saldo_inicial__cliente_id=id_cliente)
    ()

    for venta in ventas_no_pagadas:
        venta.saldo = venta.getSaldo()

    context = {
        "cliente": cliente,
        "ventas": ventas_no_pagadas,
        "abonosSI": abonos,
    }
    return render(request, "clientes/detalles_cliente.html", context=context)


@login_required(login_url="/accounts/login/")
def VistaAgregarCliente(request):
    context={
        "rutas": Ruta.objects.filter(activo=True),
    }
    if request.method == 'POST':
        try:
            data = request.POST
            saldo = Decimal(request.POST.get('saldo', 0))
            attributes = {
                "nombre": data['nombre'],
                "direccion": data['direccion'],
                "telefono": data['telefono'],
                "saldo": saldo,
            }
            ruta_id = data['ruta_select']
            ruta_instance = Ruta.objects.get(id=ruta_id)
            attributes["ruta"] = ruta_instance

            if Cliente.objects.filter(**attributes).exists():
                messages.error(request, '¡El cliente ya existe!',
                    extra_tags="warning")
                return redirect('Apps.clientes:agregar_cliente')
            with transaction.atomic():
                new_customer = Cliente.objects.create(**attributes)

                idclinete= new_customer.id
                if saldo <= 0:
                    attributesSaldo = {
                        "cliente": new_customer,
                        "saldo_inicial": saldo,
                        "estado": 0,
                    } 
                else:
                    attributesSaldo = {
                        "cliente": new_customer,
                        "saldo_inicial": saldo,
                    }

                new_saldo = SaldoInicial.objects.create(**attributesSaldo)

                messages.success(request, '¡Cliente: ' + attributes["nombre"] + " " + ' Creado con éxito!', extra_tags="success")
                return redirect('Apps.clientes:lista_clientes')
        except Exception as e:
            messages.success(request, f'Error en cliente: {str(e)}', extra_tags="danger")
            return redirect('Apps.clientes:agregar_cliente')
    return render(request, "clientes/agregar_clientes.html",context=context)


@login_required(login_url="/accounts/login/")
def VistaActulizarCliente(request, id_cliente):
    try:
        cliente = Cliente.objects.get(id=id_cliente)
    except Exception as e:
        messages.success(
            request, '¡Hubo un error al intentar localizar al cliente!', extra_tags="danger")
        return redirect('Apps.clientes:lista_clientes')
    context = {
        "cliente": cliente,
    }
    if request.method == 'POST':
        try:
            data = request.POST
            attributes = {
                "nombre": data['nombre'],
                "direccion": data['direccion'],
                "telefono": data['telefono'],
            }
            ruta_id = data['ruta_select']
            ruta_instance = Ruta.objects.get(id=ruta_id)
            attributes["ruta"] = ruta_instance
            cliente = Cliente.objects.filter(
                id=id_cliente).update(**attributes)
            cliente = Cliente.objects.get(id=id_cliente)
            messages.success(request, '¡Cliente: ' + cliente.nombre +
                ' actualizado exitosamente!', extra_tags="success")
            return redirect('Apps.clientes:lista_clientes')
        except Exception as e:
            messages.success(
                request, '¡Hubo un error durante la actualización!', extra_tags="danger")
            return redirect('Apps.clientes:lista_clientes')
    return render(request, "clientes/actualizar_ clientes.html", context=context)


@login_required(login_url="/accounts/login/")
def VistaEliminarCliente(request, id_cliente):
    try:
        cliente = Cliente.objects.get(id=id_cliente)
        cliente.activo = False
        cliente.save()
        messages.success(request, '¡Cliente: ' + cliente.nombre +
            ' Eliminado!', extra_tags="success")
        return redirect('Apps.clientes:lista_clientes')
    except Exception as e:
        messages.success(
            request, '¡Hubo un error durante la eliminación!', extra_tags="danger")
        return redirect('Apps.clientes:lista_clientes')
    
@login_required(login_url="/accounts/login/")
def VistaPagarSaldo(request,id_cliente):
    try:
        fecha_actual = timezone.now().strftime('%Y-%m-%d')
        saldo_inicial = SaldoInicial.objects.get(cliente_id=id_cliente)
        monto= saldo_inicial.saldo_inicial - saldo_inicial.abono_saldo
        if saldo_inicial.estado:
            
            AbonoSaldoInicial.objects.create(
                id_saldo_inicial=saldo_inicial,
                fecha=fecha_actual,
                monto_SI=monto
            )
            saldo_inicial.abono_saldo = saldo_inicial.saldo_inicial
            saldo_inicial.estado = False
            saldo_inicial.save()

            saldo_total_cliente = Cliente.objects.get(pk=id_cliente)
            saldo_total_cliente.saldo -= monto
            saldo_total_cliente.save()
        
        messages.success(request, '¡Cliente: ' + saldo_total_cliente.nombre +
            ' ha saldado su deuda inicial!', extra_tags="success")
        return DetalleClientesView(request, id_cliente)
    except Exception as e:
        messages.success(
            request, '¡Hubo un error durante el abono!', extra_tags="danger")
        return DetalleClientesView(request, id_cliente)
    
def is_ajax(request):
    return request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest'

def GetClientsAJAXView(request):
    if request.method == 'POST':
        if is_ajax(request=request):
            data = []
            clientes = Cliente.objects.filter(nombre__icontains=request.POST['term'],activo=True)
            for cliente in clientes[0:10]:
                item = cliente.to_json()
                data.append(item)

            return JsonResponse(data, safe=False)
        


#para reportes---------------
def VistaCliente(request):
    clientes = Cliente.objects.filter(activo=True)
    print(clientes)  # Añadir esta línea para imprimir los clientes en la consola
    context = {
        "clientes": clientes,
    }

    return render(request, "ventas/Vreport.html", context=context)







@login_required(login_url="/accounts/login/")
def generar_reporte_clientes_excel(request):
    clientes = Cliente.objects.filter(saldo__gt=0, activo=True)

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=reporte_clientes.xlsx'

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Reporte de Clientes"

    headers = ['Nombre', 'Dirección', 'Teléfono', 'Saldo']

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = worksheet.cell(row=2, column=col_num, value=header)

        cell.fill = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid')
        cell.border = Border(left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'),
                            top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'))
        cell.alignment = Alignment(horizontal='center')

    for row_num, cliente in enumerate(clientes, 3):
        worksheet.cell(row=row_num, column=1, value=cliente.nombre).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=2, value=cliente.direccion).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=3, value=cliente.telefono).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=4, value=float(cliente.saldo)).alignment = Alignment(horizontal='center')

    for col_num in range(1, worksheet.max_column + 1):
        max_length = 0
        column = get_column_letter(col_num)
        for cell in worksheet[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width

    workbook.save(response)

    return response




@login_required(login_url="/accounts/login/")
def generar_reporte_clientes_pdf(request):
    clientes = Cliente.objects.filter(saldo__gt=0, activo=True)

    # Creamos un objeto HttpResponse con tipo de contenido para PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=reporte_clientes.pdf'

    # Creamos un objeto BytesIO para almacenar el PDF
    buffer = BytesIO()

    # Creamos un documento PDF
    pdf = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    style_title = styles['Title']
    style_table_header = styles['Heading4']
    style_table_cell = ParagraphStyle(
        'TableCellStyle', parent=styles['Normal'], alignment=TA_CENTER)

    # Lista de datos para la tabla
    data = []
    data.append(['Nombre', 'Dirección', 'Teléfono', 'Saldo'])

    # Recorremos los clientes y agregamos los datos a la lista
    for cliente in clientes:
        data.append([
            cliente.nombre,
            cliente.direccion,
            cliente.telefono,
            str(cliente.saldo),
        ])

    # Creamos la tabla y le aplicamos estilo
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.gray),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
    ]))

    # Construimos el documento PDF
    elements = []
    elements.append(table)
    pdf.build(elements)

    # Obtenemos el valor del buffer y lo escribimos en la respuesta HTTP
    pdf_data = buffer.getvalue()
    buffer.close()
    response.write(pdf_data)

    return response



@login_required(login_url="/accounts/login/")
def obtener_datos_filtrados_clientes(request):
    clientes = Cliente.objects.filter(saldo__gt=0)

    data = []

    for cliente in clientes:
        data.append({
            'Nombre': cliente.nombre,
            'Direccion': cliente.direccion,
            'Telefono': cliente.telefono,
            'Saldo': cliente.saldo,
        })

    return JsonResponse(data, safe=False)



@login_required(login_url="/accounts/login/")
def ClienteView(request):

    return render(request, "clientes/Clreport.html")


