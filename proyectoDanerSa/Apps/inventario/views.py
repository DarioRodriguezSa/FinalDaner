from django.shortcuts import render, get_object_or_404, redirect
from .models import Producto
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from decimal import Decimal
from django.shortcuts import redirect
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_protect
from decimal import Decimal, InvalidOperation
from django.http import HttpResponseBadRequest


from openpyxl.styles import  PatternFill, Border, Side
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


#Para PDF
from openpyxl.styles import PatternFill, Border, Side, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from io import BytesIO



@login_required(login_url="/accounts/login/")
def index(request):
    productos = Producto.objects.all()
    productos_a_vencer = Producto.objects.filter(existencia__lte=5)
    return render(request, "inventario/index.html", {'productos': productos, 'productos_a_vencer': productos_a_vencer})

@login_required(login_url="/accounts/login/")
@csrf_protect
def eliminar_producto(request, producto_id):
    producto = get_object_or_404(Producto, idproducto=producto_id)

    if request.method == 'POST' and producto.estado != Producto.INACTIVO:
        # Cambia el estado del producto a inactivo
        producto.estado = Producto.INACTIVO
        producto.save()
        messages.success(request, '')
        return JsonResponse({'success': True, 'message': 'El producto fue marcado como inactivo.'})

    return JsonResponse({'success': False, 'message': 'No se pudo eliminar el producto.'})

@login_required(login_url="/accounts/login/")
def modificar_producto(request, producto_id):
    producto = get_object_or_404(Producto, idproducto=producto_id)

    if request.method == 'POST':
        producto.nombre = request.POST.get('nombre', producto.nombre)

        existencia = request.POST.get('existencia', '')
        try:
            existencia = float(existencia)
            if existencia < 0:
                raise ValueError("La existencia no puede ser un número negativo.")
            producto.existencia = existencia
        except ValueError:
            return HttpResponseBadRequest("Error: La existencia debe ser un número válido y no puede ser negativa.")

        precio_venta = request.POST.get('precio_venta', '')
        try:
            precio_venta = Decimal(precio_venta)
            if precio_venta < 0:
                raise ValueError("El precio de venta no puede ser un número negativo.")
            producto.precio_venta = precio_venta
        except InvalidOperation:
            return HttpResponseBadRequest("Error: El precio de venta debe ser un número válido y no puede ser negativo.")

        precio_compra = request.POST.get('precio_compra', '')
        try:
            precio_compra = Decimal(precio_compra)
            if precio_compra < 0:
                raise ValueError("El precio de compra no puede ser un número negativo.")
            if precio_compra > producto.precio_venta:
                raise ValueError("El precio de compra no puede ser mayor que el precio de venta.")
            producto.precio_compra = precio_compra
        except InvalidOperation:
            return HttpResponseBadRequest("Error: El precio de compra debe ser un número válido y no puede ser negativo.")

        producto.estado = int(request.POST.get('estado', producto.estado))

        producto.save()

        # Aquí defines los mensajes de éxito o error según sea necesario
        mensaje = 'Producto modificado exitosamente'
        # Si hay algún error en la validación, puedes definir los mensajes de error aquí
        mensaje_nombre = "Mensaje de error para el nombre"
        mensaje_existencia = "Mensaje de error para la existencia"
        mensaje_precio_venta = "Mensaje de error para el precio de venta"
        mensaje_precio_compra = "Mensaje de error para el precio de compra"

        return render(request, "inventario/index.html", {'producto': producto,
                                                        'mensaje_nombre': mensaje_nombre,
                                                        'mensaje_existencia': mensaje_existencia,
                                                        'mensaje_precio_venta': mensaje_precio_venta,
                                                        'mensaje_precio_compra': mensaje_precio_compra})


    return render(request, "inventario/index.html", {'producto': producto})

@login_required(login_url="/accounts/login/")
def VistaAgregarProducto(request):
    if request.method == 'POST':
        data = request.POST
        nombre = data['nombre']
        existencia = float(data['existencia'])
        precio_venta = Decimal(data['precio_venta'])
        precio_compra = Decimal(data['precio_compra'])
        estado = int(data['estado'])
        
        if precio_compra >= precio_venta:
            messages.error(request, '¡El precio de compra debe ser menor que el precio de venta!.', extra_tags="warning")
            return redirect('Apps.inventario:agregar_producto')
        
        if Producto.objects.filter(nombre=nombre).exists():
            messages.error(request, '¡El producto ya existe!', extra_tags="warning")
            return redirect('Apps.inventario:agregar_producto')
        
        try:
            new_product = Producto.objects.create(
                nombre=nombre,
                existencia=existencia,
                precio_venta=precio_venta,
                precio_compra=precio_compra,
                estado=estado
            )
            messages.success(request, '¡Producto: ' + nombre + ' creado con éxito!', extra_tags="success")
            return redirect('Apps.inventario:index')
        except Exception as e:
            messages.error(request, '¡Hubo un error durante la creación!', extra_tags="danger")
            print(e)
            return redirect('Apps.inventario:agregar_producto')
    
    return render(request, "inventario/agregar_producto.html")

@login_required(login_url="/accounts/login/")
def InventarioView(request):

    return render(request, "inventario/Ireport.html")      

from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from .models import Producto

@login_required(login_url="/accounts/login/")
def generar_reporte_inventario_excel(request):
    productos = Producto.objects.all()

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=reporte_inventario.xlsx'

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Reporte de Inventario"

    headers = ['Nombre', 'Existencia', 'Precio de Compra', 'Precio de Venta','Total Compras', 'Total Ventas', 'Estado']

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = worksheet.cell(row=2, column=col_num, value=header)

        cell.fill = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid')
        cell.border = Border(left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'),
                            top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'))
        cell.alignment = Alignment(horizontal='center')

    for row_num, producto in enumerate(productos, 3):
        total_compras = producto.existencia * producto.precio_compra
        total_ventas = producto.existencia * producto.precio_venta

        worksheet.cell(row=row_num, column=1, value=producto.nombre).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=2, value=producto.existencia).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=3, value=producto.precio_compra).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=4, value=producto.precio_venta).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=7, value=producto.get_estado_display()).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=5, value=total_compras).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=6, value=total_ventas).alignment = Alignment(horizontal='center')

    for col_num in range(1, worksheet.max_column + 1):
        max_length = 0
        column = get_column_letter(col_num)
        for cell in worksheet[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width

    workbook.save(response)

    return response

@login_required(login_url="/accounts/login/")
def generar_reporte_inventario_pdf(request):
    productos = Producto.objects.all()

    # Creamos un objeto HttpResponse con tipo de contenido para PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=reporte_inventario.pdf'

    # Creamos un objeto BytesIO para almacenar el PDF
    buffer = BytesIO()

    # Creamos un documento PDF
    pdf = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    style_title = styles['Title']
    style_table_header = styles['Heading4']
    style_table_cell = ParagraphStyle(
        'TableCellStyle', parent=styles['Normal'], alignment=TA_CENTER)

    # Lista de datos para la tabla
    data = []
    data.append(['Nombre', 'Existencia', 'Precio de Compra', 'Precio de Venta', 'Estado', 'Total Compras', 'Total Ventas'])

    # Recorremos los productos y agregamos los datos a la lista
    for producto in productos:
        total_compras = producto.existencia * producto.precio_compra
        total_ventas = producto.existencia * producto.precio_venta
        data.append([
            producto.nombre,
            str(producto.existencia),
            str(producto.precio_compra),
            str(producto.precio_venta),
            producto.get_estado_display(),
            str(total_compras),
            str(total_ventas)
        ])

    # Creamos la tabla y le aplicamos estilo
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.gray),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
    ]))

    # Construimos el documento PDF
    elements = []
    elements.append(table)
    pdf.build(elements)

    # Obtenemos el valor del buffer y lo escribimos en la respuesta HTTP
    pdf_data = buffer.getvalue()
    buffer.close()
    response.write(pdf_data)

    return response

@login_required(login_url="/accounts/login/")
def obtener_datos_filtrados_inventario(request):
    productos = Producto.objects.all()

    data = []

    for producto in productos:
        total_compras = producto.existencia * producto.precio_compra
        total_ventas = producto.existencia * producto.precio_venta

        data.append({
            'Nombre': producto.nombre,
            'Existencia': producto.existencia,
            'PrecioVenta': producto.precio_venta,
            'PrecioCompra': producto.precio_compra,
            'TotalCompras': total_compras,
            'TotalVentas': total_ventas,
            'Estado': 'Activo' if producto.estado == Producto.ACTIVO else 'Inactivo',
        })

    return JsonResponse(data, safe=False)










