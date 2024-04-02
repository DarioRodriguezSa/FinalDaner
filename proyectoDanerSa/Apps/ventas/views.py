from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from Apps.inventario.models import Producto
from Apps.clientes.models import Cliente, SaldoInicial, AbonoSaldoInicial
from .models import Venta,DetalleVenta,Transaccion
from django.contrib import messages
from django.db import transaction
from decimal import Decimal, InvalidOperation
from django.shortcuts import get_object_or_404
from django.core.exceptions import ObjectDoesNotExist
from datetime import timedelta,datetime



#------Importaciones para reportes-Excel///
from django.utils import timezone 
from openpyxl.styles import Font, PatternFill, Border, Side
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from django.db.models import Sum
from django.http import JsonResponse


#------Importaciones para reportes-PDF///
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from .models import Venta, DetalleVenta, Transaccion
from django.utils import timezone
from django.db.models import Sum
from reportlab.lib.units import cm, inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from django.utils.html import escape
from django.db.models import Sum
from django.shortcuts import render, redirect
from reportlab.lib.enums import TA_CENTER
from django.utils.dateparse import parse_date
from django.core.exceptions import ValidationError





@login_required(login_url="/accounts/login/")
def ListaVentasView(request):
    ventas = Venta.objects.all()
    context = {
        "ventas": ventas,
    }
    return render(request, "ventas/ventas.html", context=context)


@login_required(login_url="/accounts/login/")
@transaction.atomic
def VistaAgregarVentas(request):
    if request.method == 'POST':
        try:
            cliente_id = request.POST.get('cliente')
            monto = Decimal(request.POST.get('anticipo', 0))
            total = Decimal(request.POST.get('total', 0))
            comentario = request.POST.get('comentario', '')
            id_fecha = request.POST.get('id_fecha')
            fecha_cobro = datetime.strptime(id_fecha, '%Y-%m-%d') + timedelta(days=7)

            with transaction.atomic():
                venta = Venta.objects.create(
                    id_clientes=Cliente.objects.get(pk=cliente_id),
                    fecha=id_fecha,
                    fecha_cobro=fecha_cobro.date(),  # Establecer la fecha de cobro calculada
                    anticipo=0,
                    total=total,
                    comentario=comentario,
                    estado_cuenta=True,
                    user=request.user,
                )

                crear_detalles_venta(request, venta)
                saldo_total_cliente = Cliente.objects.get(pk=cliente_id)
                saldo_total_cliente.saldo += total
                saldo_total_cliente.save()

                saldo_inicial = SaldoInicial.objects.get(cliente=cliente_id)
                if monto > 0:
                    if saldo_inicial.estado:
                        if monto <= saldo_inicial.saldo_inicial - saldo_inicial.abono_saldo:
                            saldo_inicial.abono_saldo += monto
                            saldo_inicial.save()
                    
                            saldo_total_cliente = Cliente.objects.get(pk=cliente_id)
                            saldo_total_cliente.saldo -= monto
                            saldo_total_cliente.save()
                            CreacionAbonoSaldoInicial(saldo_inicial.pk, id_fecha, monto)

                            if saldo_inicial.saldo_inicial == saldo_inicial.abono_saldo:
                                saldo_inicial.estado = False
                                saldo_inicial.save()
                        else:
                            abono_saldo_ultimo=  saldo_inicial.saldo_inicial - saldo_inicial.abono_saldo
                            monto_excedente = monto - abono_saldo_ultimo
                            CreacionAbonoSaldoInicial(saldo_inicial.pk, id_fecha, abono_saldo_ultimo)
                            saldo_inicial.abono_saldo = saldo_inicial.saldo_inicial
                            saldo_inicial.estado = False
                            saldo_inicial.save()
                            

                            saldo_total_cliente = Cliente.objects.get(pk=cliente_id)
                            saldo_total_cliente.saldo -= abono_saldo_ultimo
                            saldo_total_cliente.save()
                            ObtenerDeudasAntiguas(cliente_id, monto_excedente, id_fecha)
                    else:
                        ObtenerDeudasAntiguas(cliente_id,monto,id_fecha)
                
                messages.success(request, '¡Venta creada con éxito!', extra_tags="success")
        except ObjectDoesNotExist:
            messages.error(request, 'El cliente no existe', extra_tags="danger")
        except InvalidOperation as e:
            messages.error(request, f'Error en operación decimal: {str(e)}', extra_tags="danger")
        except Exception as e:
            messages.error(request, f'Error al procesar la venta: {str(e)}', extra_tags="danger")
        return redirect('Apps.ventas:lista_ventas')

    productos = Producto.objects.all()
    clientes = Cliente.objects.all()
    context = {
        "productos": productos,
        "clientes": clientes,
    }
    return render(request, "ventas/agregar_ventas.html", context=context)

@login_required(login_url="/accounts/login/")
@transaction.atomic
def VistaActulizarVenta(request, id_venta):
    try:
        ventas = Venta.objects.get(id=id_venta)
    except Exception as e:
        messages.success(
            request, '¡Hubo un error al intentar localizar la venta!', extra_tags="danger")
        return redirect('Apps.ventas:lista_ventas')
    context = {
        "ventas": ventas,
    }
    if request.method == 'POST':
        try:
            data = request.POST
            attributes = {
                "fecha": data['fecha'],
                "fecha_cobro": data['fecha_cobro'],
                "comentario": data['comentario'],
            }

            venta = Venta.objects.filter(
                id=id_venta).update(**attributes)
            venta = Venta.objects.get(id=id_venta)
            messages.success(request, '¡Venta actualizado exitosamente!', extra_tags="success")
            return redirect('Apps.ventas:lista_ventas')
        except Exception as e:
            messages.success(
                request, '¡Hubo un error durante la actualización!', extra_tags="danger")
            return redirect('Apps.ventas:lista_ventas')
    return render(request, "clientes/actualizar_ clientes.html", context=context)

@login_required(login_url="/accounts/login/")
@transaction.atomic
def VistaEliminarVenta(request, id_venta):
    try:
        venta = Venta.objects.get(id=id_venta)
        cliente = venta.id_clientes
        total_transaccion =  calcular_total_transacciones(id_venta)

        total_pagado= total_transaccion + venta.anticipo
        total_pendiente = venta.total - total_pagado

        cliente.saldo -= total_pendiente
        cliente.save()

        reintegrar_productos_inventario(id_venta)
        
        Transaccion.objects.filter(id_venta_id=id_venta).delete()
        DetalleVenta.objects.filter(id_venta_id=id_venta).delete()
        venta.delete()

        messages.success(request, '¡Venta eliminada exitosamente!', extra_tags="success")
        return redirect('Apps.ventas:lista_ventas')
    except Exception as e:
            messages.error(request, 'Ha ocurrido un error al intentar eliminar la venta', extra_tags="danger")
            return redirect('Apps.ventas:lista_ventas')

@login_required(login_url="/accounts/login/")
def crear_detalles_venta(request, venta):
    productos_ids = request.POST.getlist('productos_ids[]')
    productos_cantidades = request.POST.getlist('productos_cantidades[]')
    productos_precios = request.POST.getlist('productos_precios[]')

    for productos, cantidades, precios in zip(productos_ids, productos_cantidades, productos_precios):
        lista_productos = productos.split(',')
        lista_cantidades = cantidades.split(',')
        lista_precios = precios.split(',')

        for producto_id, cantidad, precio in zip(lista_productos, lista_cantidades, lista_precios):
            producto = Producto.objects.get(idproducto=producto_id)
            cantidad_vendida = int(cantidad)

            if producto.existencia >= cantidad_vendida:
                producto.existencia -= cantidad_vendida
                producto.save()
            else:
                raise Exception('No hay suficiente existencia para realizar la venta.')
            ganancia = Decimal(precio) - producto.precio_compra 

            DetalleVenta.objects.create(
                id_venta=venta,
                id_producto=producto,
                cantidad=cantidad,
                precio=precio,
                ganancia=ganancia
            )

def calcular_total_transacciones(id_venta):
    try:
        venta = Venta.objects.get(id=id_venta)
        cliente = venta.id_clientes
        total_transaccion = 0

        transacciones = Transaccion.objects.filter(id_venta_id=id_venta)

        for transaccion in transacciones:
            total_transaccion += transaccion.abono
        return total_transaccion

    except Venta.DoesNotExist:
        return 0

    except Transaccion.DoesNotExist:
        return 0

def reintegrar_productos_inventario(id_venta):
    detalles_venta = DetalleVenta.objects.filter(id_venta_id=id_venta)

    for detalle_venta in detalles_venta:
        producto = detalle_venta.id_producto
        cantidad = detalle_venta.cantidad

        producto.existencia += cantidad
        producto.save()

def CreacionAbonoSaldoInicial(id_SI, fecha, monto):
    saldo_inicial = SaldoInicial.objects.get(pk=id_SI)
    AbonoSaldoInicial.objects.create(
        id_saldo_inicial=saldo_inicial,
        fecha=fecha,
        monto_SI=monto,
    )

def ObtenerDeudasAntiguas(cliente_id, monto, fecha):
    ventas_no_pagadas = Venta.objects.filter(id_clientes=cliente_id, estado_cuenta=True).order_by('fecha').first()
    ultima_transaccion = Transaccion.objects.filter(id_venta=ventas_no_pagadas.id).order_by('-id').first()
    
    if ultima_transaccion is None:
        saldo_a_abonar = min(monto, ventas_no_pagadas.getSaldo())
        if isinstance(fecha, str):
            fecha = datetime.strptime(fecha, '%Y-%m-%d').date()
        if isinstance(ventas_no_pagadas.fecha, str):
            ventas_no_pagadas.fecha = datetime.strptime(ventas_no_pagadas.fecha, '%Y-%m-%d').date()

        if  fecha == ventas_no_pagadas.fecha:
            AbonarAnticipo(ventas_no_pagadas.id, saldo_a_abonar)
        else:
            AbonarTransaccion(ventas_no_pagadas.id, saldo_a_abonar)

        if saldo_a_abonar < monto:
            ObtenerDeudasAntiguas(cliente_id, monto - saldo_a_abonar, fecha)
    else:
        saldo_a_abonar = min(monto, ultima_transaccion.saldo_venta)
        AbonarTransaccion(ventas_no_pagadas.id, saldo_a_abonar)

        if saldo_a_abonar < monto:
            ObtenerDeudasAntiguas(cliente_id, monto - saldo_a_abonar, fecha)


def AbonarTransaccion(id_venta, abono):
    venta = get_object_or_404(Venta, pk=id_venta)
    abono = abono
    cliente = venta.id_clientes
    with transaction.atomic():
            saldo_venta = Decimal('0.0') if venta.getSaldo() is None else venta.getSaldo()
            saldo_cliente = Decimal('0.0') if cliente.saldo is None else cliente.saldo

            cliente.saldo = saldo_cliente - abono
            cliente.save()

            if saldo_venta - abono == 0:
                venta.estado_cuenta = False
                venta.save()
                Transaccion.objects.create(
                id_cliente=cliente,
                id_venta=venta,
                abono=abono,
                saldo_venta=saldo_venta - abono,
                saldo_cliente=saldo_cliente - abono
                )
            else:
                ultima_transaccion = Transaccion.objects.filter(id_venta=id_venta).order_by('-id').first()
                if ultima_transaccion:
                    restante = ultima_transaccion.saldo_venta - abono
                    if restante == 0:
                        venta.estado_cuenta = False
                        venta.save()
                        Transaccion.objects.create(
                            id_cliente=cliente,
                            id_venta=venta,
                            abono=abono,
                            saldo_venta=ultima_transaccion.saldo_venta - abono,
                            saldo_cliente=saldo_cliente - abono
                        )
                    else:
                            Transaccion.objects.create(
                            id_cliente=cliente,
                            id_venta=venta,
                            abono=abono,
                            saldo_venta=ultima_transaccion.saldo_venta - abono,
                            saldo_cliente=saldo_cliente - abono
                        )
                else: 
                        Transaccion.objects.create(
                            id_cliente=cliente,
                            id_venta=venta,
                            abono=abono,
                            saldo_venta=saldo_venta - abono,
                            saldo_cliente=saldo_cliente - abono
                        )

def AbonarAnticipo(id_venta, anticipo):
    venta = get_object_or_404(Venta, pk=id_venta)
    abono = anticipo
    cliente = venta.id_clientes

    cliente.saldo = cliente.saldo - abono
    cliente.save()

    venta.anticipo = abono
    if venta.anticipo == venta.total:
        venta.estado_cuenta = False
    venta.save()
            

@login_required(login_url="/accounts/login/")
def AbonarTransaccion2(request, id_venta):
    venta = get_object_or_404(Venta, pk=id_venta)
    abono = Decimal(request.POST.get('abono',0))
    cliente = venta.id_clientes
    try:
        with transaction.atomic():
            saldo_venta = Decimal('0.0') if venta.getSaldo() is None else venta.getSaldo()
            saldo_cliente = Decimal('0.0') if cliente.saldo is None else cliente.saldo

            cliente.saldo = saldo_cliente - abono
            cliente.save()

            if saldo_venta - abono == 0:
                venta.estado_cuenta = False
                venta.save()
                Transaccion.objects.create(
                id_cliente=cliente,
                id_venta=venta,
                abono=abono,
                saldo_venta=saldo_venta - abono,
                saldo_cliente=saldo_cliente - abono
                )
            else:
                ultima_transaccion = Transaccion.objects.filter(id_venta=id_venta).order_by('-id').first()
                if ultima_transaccion:
                    restante = ultima_transaccion.saldo_venta - abono
                    if restante == 0:
                        venta.estado_cuenta = False
                        venta.save()
                        Transaccion.objects.create(
                            id_cliente=cliente,
                            id_venta=venta,
                            abono=abono,
                            saldo_venta=ultima_transaccion.saldo_venta - abono,
                            saldo_cliente=saldo_cliente - abono
                        )
                    else:
                            Transaccion.objects.create(
                            id_cliente=cliente,
                            id_venta=venta,
                            abono=abono,
                            saldo_venta=ultima_transaccion.saldo_venta - abono,
                            saldo_cliente=saldo_cliente - abono
                        )
                else: 
                        Transaccion.objects.create(
                            id_cliente=cliente,
                            id_venta=venta,
                            abono=abono,
                            saldo_venta=saldo_venta - abono,
                            saldo_cliente=saldo_cliente - abono
                        )
            messages.success(request, '¡El abono fue un éxito!', extra_tags="success")

    except Exception as e:
        messages.error(request, f'Error al procesar el abono: {str(e)}', extra_tags="danger")

    return redirect('Apps.clientes:detalles_clientes',id_cliente=cliente.id )

#------------------------------------Reportes para ventas --------------------------------------
            

#---lleva a pagina de reportes     
@login_required(login_url="/accounts/login/")
def VentasView(request):

    return render(request, "ventas/Vreport.html")        



@login_required(login_url="/accounts/login/")
def ClientesView(request):

    return render(request, "ventas/Vreport_c.html")      


@login_required(login_url="/accounts/login/")
def ClientesCTView(request):

    return render(request, "ventas/Vreport_ct.html")  



@login_required(login_url="/accounts/login/")
def generar_reporte_excel(request, start_date, end_date):
    start_date = timezone.datetime.strptime(start_date, '%Y-%m-%d')
    end_date = timezone.datetime.strptime(end_date, '%Y-%m-%d')

    ventas = Venta.objects.filter(fecha__range=[start_date, end_date])

    if not ventas:
        # Si no hay ventas en el rango de fechas especificado, regresa una respuesta vacía o un mensaje de error
            messages.success(
            request, '¡No hay ventas al contado En la Fecha seleccionada!', extra_tags="danger")
            return redirect('Apps.ventas:report')

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=reporte_ventas.xlsx'

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Reporte de Ventas"

    headers = ['Usuario', 'Fecha', 'Cliente', 'Anticipo', 'Productos Vendidos', 'Cantidad Vendida', 'Precio de Compra', 'Precio de Venta', 'Ganancia', 'Total']

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = worksheet.cell(row=2, column=col_num, value=header)

        cell.fill = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid')
        cell.border = Border(left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'),
                            top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'))
        cell.alignment = Alignment(horizontal='center')

    total_ventas = 0  # Variable para almacenar la suma de todos los totales
    total_ganancia_general = 0

    for row_num, venta in enumerate(ventas, 3):
        detalles_venta = DetalleVenta.objects.filter(id_venta=venta)

        user_info = f"{venta.user.username}"
        worksheet.cell(row=row_num, column=1, value=user_info).alignment = Alignment(horizontal='left', wrap_text=True)

        worksheet.cell(row=row_num, column=2, value=venta.fecha).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=3, value=venta.id_clientes.nombre).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=10, value=venta.total).alignment = Alignment(horizontal='center')

        total_ventas += venta.total

        total_cantidad = detalles_venta.aggregate(Sum('cantidad'))['cantidad__sum']
        total_ganancia = detalles_venta.aggregate(Sum('ganancia'))['ganancia__sum']
        total_ganancia_general += total_ganancia  # Agregar el total de esta venta al total general

        worksheet.cell(row=row_num, column=4, value=venta.anticipo).alignment = Alignment(horizontal='center')

        productos_vendidos = ', '.join([detalle.id_producto.nombre for detalle in detalles_venta])
        worksheet.cell(row=row_num, column=5, value=productos_vendidos).alignment = Alignment(horizontal='center')

        total_cantidad = detalles_venta.aggregate(Sum('cantidad'))['cantidad__sum']
        total_ganancia = detalles_venta.aggregate(Sum('ganancia'))['ganancia__sum']

        worksheet.cell(row=row_num, column=6, value=total_cantidad).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=9, value=total_ganancia).alignment = Alignment(horizontal='center')
        

        # Nuevas líneas para obtener el precio de compra y precio de venta
        precio_venta = detalles_venta.aggregate(Sum('precio'))['precio__sum']
        ganancia_total = detalles_venta.aggregate(Sum('ganancia'))['ganancia__sum']
        precio_compra = precio_venta - ganancia_total

        worksheet.cell(row=row_num, column=7, value=precio_compra).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=8, value=precio_venta).alignment = Alignment(horizontal='center')

        ultima_transaccion = Transaccion.objects.filter(id_venta=venta).order_by('-fecha').first()
        saldo_restante = ultima_transaccion.saldo_venta if ultima_transaccion else 0

    # Agregar la suma total en la columna 10
    worksheet.cell(row=row_num + 1, column=10, value=total_ventas).alignment = Alignment(horizontal='center')
    worksheet.cell(row=row_num + 1, column=10).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    worksheet.cell(row=row_num + 1, column=9, value=total_ganancia_general).alignment = Alignment(horizontal='center')
    worksheet.cell(row=row_num + 1, column=9).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    for col_num in range(1, worksheet.max_column + 1):
        max_length = 0
        column = get_column_letter(col_num)
        for cell in worksheet[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width

    workbook.save(response)

    return response




@login_required(login_url="/accounts/login/")
def generar_reporte_pdf(request, start_date, end_date):
    # Convertir las fechas de inicio y fin de la cadena a objetos de fecha y hora
    start_date = timezone.datetime.strptime(start_date, '%Y-%m-%d')
    end_date = timezone.datetime.strptime(end_date, '%Y-%m-%d')

    # Filtrar las ventas en el rango de fechas proporcionado
    ventas = Venta.objects.filter(fecha__range=[start_date, end_date])

    # Crear un objeto de respuesta HTTP con el tipo de contenido apropiado para un PDF
    response = HttpResponse(content_type='application/pdf')
    # Establecer el encabezado Content-Disposition para que el navegador abra el PDF en línea o lo descargue como un archivo adjunto
    response['Content-Disposition'] = f'attachment; filename="reporte_ventas.pdf"'

    # Crear un documento PDF con el tamaño de página obtenido
    pdf = SimpleDocTemplate(response, pagesize=letter)
    # Crear una lista para almacenar los elementos del PDF
    elements = []

    # Crear la tabla de datos para las ventas
    data = []
    # Encabezados de la tabla
    data.append(['Usuario', 'Fecha', 'Cliente', 'Total', 'Anticipo', 'Productos Vendidos', 'Cantidad Vendida', 'Precio de Compra', 'Precio de Venta', 'Ganancia'])

    for venta in ventas:
        # Obtener detalles de la venta
        detalles_venta = DetalleVenta.objects.filter(id_venta=venta)

        # Obtener información del usuario
        user_info = venta.user.username

        # Obtener el nombre del cliente
        cliente_nombre = venta.id_clientes.nombre

        # Obtener los productos vendidos
        productos_vendidos = ', '.join([detalle.id_producto.nombre for detalle in detalles_venta])

        # Calcular el total de la cantidad vendida y la ganancia
        total_cantidad = detalles_venta.aggregate(Sum('cantidad'))['cantidad__sum']
        total_ganancia = detalles_venta.aggregate(Sum('ganancia'))['ganancia__sum']

        # Calcular el precio de compra y el precio de venta
        precio_venta = detalles_venta.aggregate(Sum('precio'))['precio__sum']
        ganancia_total = detalles_venta.aggregate(Sum('ganancia'))['ganancia__sum']
        precio_compra = precio_venta - ganancia_total

        # Obtener la última transacción para calcular el saldo restante
        ultima_transaccion = Transaccion.objects.filter(id_venta=venta).order_by('-fecha').first()

        # Agregar una fila de datos a la tabla
        data.append([user_info, str(venta.fecha), cliente_nombre, str(venta.total), str(venta.anticipo), productos_vendidos, str(total_cantidad), str(precio_compra), str(precio_venta), str(total_ganancia)])

    # Calcular la suma total de todos los totales
    total_ventas = sum(venta.total for venta in ventas)

    # Agregar la suma total al final de la columna 'Total' en la lista de datos
    data.append(['', '', '', f'Total de ventas: {total_ventas}', '', '', '', '', '', ''])

    # Establecer el ancho de las columnas individualmente
    column_widths = [1.5 * inch, 1 * inch, 1.5 * inch, 1 * inch, 1 * inch, 1.5 * inch, 1 * inch, 1 * inch, 1 * inch, 1 * inch]

    # Crear la tabla y agregar los datos
    tabla_data = []
    for row in data:
        tabla_row = []
        for item in row:
            # Escapar caracteres especiales para evitar errores en la generación del PDF
            item = escape(item)
            # Crear un objeto Paragraph para permitir el ajuste del texto
            tabla_row.append(Paragraph(item, getSampleStyleSheet()['Normal']))
        tabla_data.append(tabla_row)

    tabla = Table(tabla_data, colWidths=column_widths)

    # Establecer estilo para la tabla
    estilo_tabla = TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.gray),
                               ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                               ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                               ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                               ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                               ('BACKGROUND', (0, 1), (-1, -2), colors.beige),  # Color de fondo para celdas de datos
                               ('GRID', (0, 0), (-1, -1), 1, colors.black)])

    tabla.setStyle(estilo_tabla)
    elements.append(tabla)

    # Construir el PDF
    pdf.build(elements)

    return response




@login_required(login_url="/accounts/login/")
def generar_reporte_cliente_excel(request, id_cliente):
    cliente = get_object_or_404(Cliente, pk=id_cliente)
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename=reporte_cliente_{cliente.nombre}.xlsx'

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = f"Reporte de {cliente.nombre}"

    title_font = Font(size=14, color='FFFFFF', bold=True)
    title_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    title_border = Border(left=Side(border_style='thin', color='000000'),
                          right=Side(border_style='thin', color='000000'),
                          top=Side(border_style='thin', color='000000'),
                          bottom=Side(border_style='thin', color='000000'))

    headers = ['Cliente', 'Fecha', 'Abono', 'Saldo Restante', 'Producto Vendido', 'Venta Total']

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = worksheet.cell(row=1, column=col_num, value=header)

        cell.fill = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid')
        cell.border = Border(left=Side(border_style='thin', color='000000'),
                             right=Side(border_style='thin', color='000000'),
                             top=Side(border_style='thin', color='000000'),
                             bottom=Side(border_style='thin', color='000000'))
        cell.alignment = Alignment(horizontal='center')

    transacciones_por_venta = Transaccion.objects.filter(id_cliente=cliente, fecha__range=[start_date, end_date]).values('id_venta').distinct()
    row_num = 2

    for transaccion_venta in transacciones_por_venta:
        id_venta = transaccion_venta['id_venta']
        transacciones_venta_actual = Transaccion.objects.filter(id_cliente=cliente, id_venta=id_venta, fecha__range=[start_date, end_date])

        for i, transaccion in enumerate(transacciones_venta_actual):
            worksheet.cell(row=row_num, column=1, value=transaccion.id_venta.id_clientes.nombre).alignment = Alignment(horizontal='center')
            worksheet.cell(row=row_num, column=2, value=transaccion.fecha).alignment = Alignment(horizontal='center')
            worksheet.cell(row=row_num, column=3, value=transaccion.abono).alignment = Alignment(horizontal='center')
            worksheet.cell(row=row_num, column=4, value=transaccion.saldo_venta).alignment = Alignment(horizontal='center')

            productos_vendidos = ', '.join(transaccion.id_venta.detalleventa_set.all().values_list('id_producto__nombre', flat=True)) 
            worksheet.cell(row=row_num, column=5, value=productos_vendidos).alignment = Alignment(horizontal='center')

            if i == 0:
                total_venta_actual = Venta.objects.filter(id=id_venta, fecha__range=[start_date, end_date]).aggregate(Sum('total'))['total__sum'] or Decimal(0)
                total_venta_cell = worksheet.cell(row=row_num, column=6, value=total_venta_actual).alignment = Alignment(horizontal='center')

            row_num += 1

        row_num += 2

    for col_num in range(1, worksheet.max_column + 1):
        max_length = 0
        column = get_column_letter(col_num)
        for cell in worksheet[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width

    workbook.save(response)
    return response

@login_required(login_url="/accounts/login/")
def generar_reporte_cliente_pdf(request, id_cliente):
    cliente = get_object_or_404(Cliente, pk=id_cliente)
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=reporte_cliente_{cliente.nombre}.pdf'

    pdf = SimpleDocTemplate(response)
    styles = getSampleStyleSheet()

    data = []
    headers = ['Cliente', 'Fecha', 'Abono', 'Saldo Restante', 'Producto Vendido', 'Venta Total']
    data.append(headers)

    transacciones_por_venta = Transaccion.objects.filter(id_cliente=cliente, fecha__range=[start_date, end_date]).order_by('id_venta', 'fecha')
    current_venta_id = None
    current_venta_total = None

    for transaccion in transacciones_por_venta:
        if transaccion.id_venta_id != current_venta_id:
            if current_venta_id is not None:
                # Agrega una fila vacía entre ventas
                data.append([''] * len(headers))
            current_venta_id = transaccion.id_venta_id
            current_venta_total = Venta.objects.filter(id=current_venta_id, fecha__range=[start_date, end_date]).aggregate(Sum('total'))['total__sum'] or Decimal(0)
        
        abono = transaccion.abono if transaccion.abono is not None else ''
        saldo_restante = transaccion.saldo_venta if transaccion.saldo_venta is not None else ''
        productos_vendidos = ', '.join(transaccion.id_venta.detalleventa_set.all().values_list('id_producto__nombre', flat=True)) if transaccion.id_venta else ''

        data.append([
            transaccion.id_venta.id_clientes.nombre if transaccion.id_venta else '',
            transaccion.fecha.strftime('%Y-%m-%d') if transaccion.fecha else '',
            abono,
            saldo_restante,
            productos_vendidos,
            current_venta_total if transaccion.fecha == transaccion.id_venta.fecha else ''  # Solo muestra el total en la última transacción de la venta
        ])

    table = Table(data)

    style = TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.gray),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)])
    table.setStyle(style)

    pdf.build([table])
    return response



@login_required(login_url="/accounts/login/")
def generar_reporte_abonos_clientes_excel(request, start_date, end_date):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename=reporte_abonos_clientes.xlsx'

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Reporte de Abonos Clientes"

    title_font = Font(size=14, color='FFFFFF', bold=True)
    title_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    title_border = Border(left=Side(border_style='thin', color='000000'),
                          right=Side(border_style='thin', color='000000'),
                          top=Side(border_style='thin', color='000000'),
                          bottom=Side(border_style='thin', color='000000'))

    headers = ['Cliente', 'Fecha', 'Abono', 'Saldo Restante', 'Producto Vendido', 'Venta Total']

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = worksheet.cell(row=1, column=col_num, value=header)

        cell.fill = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid')
        cell.border = Border(left=Side(border_style='thin', color='000000'),
                             right=Side(border_style='thin', color='000000'),
                             top=Side(border_style='thin', color='000000'),
                             bottom=Side(border_style='thin', color='000000'))
        cell.alignment = Alignment(horizontal='center')

    transacciones_por_venta = Transaccion.objects.filter(fecha__range=[start_date, end_date]).values('id_venta').distinct()
    row_num = 2

    for transaccion_venta in transacciones_por_venta:
        id_venta = transaccion_venta['id_venta']
        transacciones_venta_actual = Transaccion.objects.filter(id_venta=id_venta, fecha__range=[start_date, end_date])

        for i, transaccion in enumerate(transacciones_venta_actual):
            worksheet.cell(row=row_num, column=1, value=transaccion.id_venta.id_clientes.nombre).alignment = Alignment(horizontal='center')
            worksheet.cell(row=row_num, column=2, value=transaccion.fecha).alignment = Alignment(horizontal='center')
            worksheet.cell(row=row_num, column=3, value=transaccion.abono).alignment = Alignment(horizontal='center')
            worksheet.cell(row=row_num, column=4, value=transaccion.saldo_venta).alignment = Alignment(horizontal='center')

            productos_vendidos = ', '.join(transaccion.id_venta.detalleventa_set.all().values_list('id_producto__nombre', flat=True)) 
            worksheet.cell(row=row_num, column=5, value=productos_vendidos).alignment = Alignment(horizontal='center')

            if i == 0:
                total_venta_actual = Venta.objects.filter(id=id_venta, fecha__range=[start_date, end_date]).aggregate(Sum('total'))['total__sum'] or Decimal(0)
                total_venta_cell = worksheet.cell(row=row_num, column=6, value=total_venta_actual).alignment = Alignment(horizontal='center')

            row_num += 1

        row_num += 2

    for col_num in range(1, worksheet.max_column + 1):
        max_length = 0
        column = get_column_letter(col_num)
        for cell in worksheet[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width

    workbook.save(response)
    return response



@login_required(login_url="/accounts/login/")
def generar_reporte_abonos_clientes_pdf(request, start_date, end_date):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_abonos_clientes.pdf"'

    # Create a PDF document
    pdf = SimpleDocTemplate(response, pagesize=letter)
    styles = getSampleStyleSheet()
    style_title = styles['Title']
    style_table_header = styles['Heading4']
    style_table_cell = ParagraphStyle(
        'TableCellStyle', parent=styles['Normal'], alignment=TA_CENTER)

    # Data for the table
    data = []
    data.append(['Cliente', 'Fecha', 'Abono', 'Saldo Restante', 'Producto Vendido', 'Venta Total'])

    transacciones_por_venta = Transaccion.objects.filter(fecha__range=[start_date, end_date]).values('id_venta').distinct()

    for transaccion_venta in transacciones_por_venta:
        id_venta = transaccion_venta['id_venta']
        transacciones_venta_actual = Transaccion.objects.filter(id_venta=id_venta, fecha__range=[start_date, end_date])

        for i, transaccion in enumerate(transacciones_venta_actual):
            productos_vendidos = ', '.join(transaccion.id_venta.detalleventa_set.all().values_list('id_producto__nombre', flat=True)) 
            data.append([
                transaccion.id_venta.id_clientes.nombre,
                str(transaccion.fecha),
                str(transaccion.abono),
                str(transaccion.saldo_venta),
                productos_vendidos,
                str(Venta.objects.filter(id=id_venta, fecha__range=[start_date, end_date]).aggregate(Sum('total'))['total__sum'] or Decimal(0))
            ])
        # Agregar una fila vacía entre cada venta
        data.append([''] * len(data[0]))

    # Create the table and style
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.gray),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
    ]))

    # Build PDF document
    elements = []
    elements.append(table)
    pdf.build(elements)

    return response









#nuevoooo
@login_required(login_url="/accounts/login/")
def obtener_datos_filtrados(request, start_date, end_date):
    start_date = timezone.datetime.strptime(start_date, '%Y-%m-%d')
    end_date = timezone.datetime.strptime(end_date, '%Y-%m-%d')

    ventas = Venta.objects.filter(fecha__range=[start_date, end_date])

    data = []

    for venta in ventas:
        detalles_venta = DetalleVenta.objects.filter(id_venta=venta)

        user_info = f"{venta.user.username}"
        fecha = venta.fecha.strftime('%Y-%m-%d')
        cliente = venta.id_clientes.nombre
        total = venta.total
        anticipo = venta.anticipo

        productos_vendidos = ', '.join([detalle.id_producto.nombre for detalle in detalles_venta])

        total_cantidad = detalles_venta.aggregate(Sum('cantidad'))['cantidad__sum']
        total_ganancia = detalles_venta.aggregate(Sum('ganancia'))['ganancia__sum']

        precio_venta = detalles_venta.aggregate(Sum('precio'))['precio__sum']
        ganancia_total = detalles_venta.aggregate(Sum('ganancia'))['ganancia__sum']
        precio_compra = precio_venta - ganancia_total

        ultima_transaccion = Transaccion.objects.filter(id_venta=venta).order_by('-fecha').first()
        saldo_restante = ultima_transaccion.saldo_venta if ultima_transaccion else 0

        data.append({
            'Usuario': user_info,
            'Fecha': fecha,
            'Cliente': cliente,
            'Total': total,
            'Anticipo': anticipo,
            'SaldoRestante': saldo_restante,
            'ProductosVendidos': productos_vendidos,
            'CantidadVendida': total_cantidad,
            'PrecioCompra': precio_compra,
            'PrecioVenta': precio_venta,
            'Ganancia': total_ganancia,
        })

    return JsonResponse(data, safe=False)



@login_required(login_url="/accounts/login/")
def obtener_datos_abonos_filtrados(request, start_date, end_date):
    start_date = timezone.datetime.strptime(start_date, '%Y-%m-%d')
    end_date = timezone.datetime.strptime(end_date, '%Y-%m-%d')

    transacciones_por_venta = Transaccion.objects.filter(fecha__range=[start_date, end_date]).values('id_venta').distinct()
    data = []

    for transaccion_venta in transacciones_por_venta:
        id_venta = transaccion_venta['id_venta']
        transacciones_venta_actual = Transaccion.objects.filter(id_venta=id_venta, fecha__range=[start_date, end_date])

        for i, transaccion in enumerate(transacciones_venta_actual):
            productos_vendidos = ', '.join(transaccion.id_venta.detalleventa_set.all().values_list('id_producto__nombre', flat=True)) 

            if i == 0:
                total_venta_actual = Venta.objects.filter(id=id_venta, fecha__range=[start_date, end_date]).aggregate(Sum('total'))['total__sum'] or Decimal(0)
                total_venta_cell = total_venta_actual
            else:
                total_venta_cell = ''

            data.append({
                'Cliente': transaccion.id_venta.id_clientes.nombre,
                'Fecha': transaccion.fecha.strftime('%Y-%m-%d'),
                'Abono': transaccion.abono,
                'SaldoRestante': transaccion.saldo_venta,
                'ProductosVendidos': productos_vendidos,
                'VentaTotal': total_venta_cell,
            })

    return JsonResponse(data, safe=False)



@login_required(login_url="/accounts/login/")
def obtener_datos_abonos_filtrados_id(request, id_cliente):
    cliente = get_object_or_404(Cliente, pk=id_cliente)
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    try:
        # Intenta convertir las fechas al formato correcto
        start_date = parse_date(start_date)
        end_date = parse_date(end_date)
    except ValidationError:
        # Manejar el error de formato de fecha
        return JsonResponse({'error': 'Formato de fecha inválido'}, status=400)

    datos_filtrados = []

    transacciones_por_venta = Transaccion.objects.filter(id_cliente=cliente, fecha__range=[start_date, end_date]).values('id_venta').distinct()

    for transaccion_venta in transacciones_por_venta:
        id_venta = transaccion_venta['id_venta']
        transacciones_venta_actual = Transaccion.objects.filter(id_cliente=cliente, id_venta=id_venta, fecha__range=[start_date, end_date])

        for i, transaccion in enumerate(transacciones_venta_actual):
            datos_filtrados.append({
                'Cliente': transaccion.id_venta.id_clientes.nombre,
                'Fecha': transaccion.fecha,
                'Abono': transaccion.abono,
                'SaldoRestante': transaccion.saldo_venta,
                'ProductosVendidos': ', '.join(transaccion.id_venta.detalleventa_set.all().values_list('id_producto__nombre', flat=True)),
                'VentaTotal': Venta.objects.filter(id=id_venta, fecha__range=[start_date, end_date]).aggregate(Sum('total'))['total__sum'] or Decimal(0),
            })

    return JsonResponse(datos_filtrados, safe=False)





from django.db.models import F, Sum

def redireccion_fcobros(request):
    if request.method == 'POST':
        fecha_seleccionada = request.POST.get('fecha_seleccionada')
        if not fecha_seleccionada:
            return render(request, 'ventas/fcobros.html', {'mensaje': 'No hay fecha seleccionada'})

        # Filtrar ventas para la fecha seleccionada
        ventas_cobrar = Venta.objects.filter(fecha_cobro=fecha_seleccionada)

        # Filtrar clientes asociados a estas ventas
        clientes_cobrar = Cliente.objects.filter(id__in=ventas_cobrar.values('id_clientes'))

        # Filtrar clientes únicos
        clientes_cobrar = clientes_cobrar.distinct()

        # Filtrar ventas donde total no es igual a anticipo
        ventas_cobrar = ventas_cobrar.exclude(total=F('anticipo'))

        # Filtrar clientes con estado_cuenta igual a 1 en las ventas asociadas
        clientes_cobrar = clientes_cobrar.filter(id__in=ventas_cobrar.filter(estado_cuenta=1).values('id_clientes'))

        return render(request, 'ventas/fcobros.html', {'clientes_cobrar': clientes_cobrar})
    else:
        return render(request, 'ventas/fcobros.html')
    



#--------------------------------- VENTAS AL CONTADO---------------------------
@login_required(login_url="/accounts/login/")
def VentasCView(request):

    return render(request, "ventas/VreportC.html")


@login_required(login_url="/accounts/login/")
def obtener_datos_filtrados_contado(request, start_date, end_date):
    start_date = timezone.datetime.strptime(start_date, '%Y-%m-%d')
    end_date = timezone.datetime.strptime(end_date, '%Y-%m-%d')

    # Filtrar las ventas donde el total y el anticipo sean iguales
    ventas = Venta.objects.filter(fecha__range=[start_date, end_date], total=F('anticipo'))

    data = []

    for venta in ventas:
        detalles_venta = DetalleVenta.objects.filter(id_venta=venta)

        user_info = f"{venta.user.username}"
        fecha = venta.fecha.strftime('%Y-%m-%d')
        cliente = venta.id_clientes.nombre
        total = venta.total
        anticipo = venta.anticipo

        productos_vendidos = ', '.join([detalle.id_producto.nombre for detalle in detalles_venta])

        total_cantidad = detalles_venta.aggregate(Sum('cantidad'))['cantidad__sum']
        total_ganancia = detalles_venta.aggregate(Sum('ganancia'))['ganancia__sum']

        precio_venta = detalles_venta.aggregate(Sum('precio'))['precio__sum']
        ganancia_total = detalles_venta.aggregate(Sum('ganancia'))['ganancia__sum']
        precio_compra = precio_venta - ganancia_total

        ultima_transaccion = Transaccion.objects.filter(id_venta=venta).order_by('-fecha').first()
        saldo_restante = ultima_transaccion.saldo_venta if ultima_transaccion else 0

        data.append({
            'Usuario': user_info,
            'Fecha': fecha,
            'Cliente': cliente,
            'Total': total,
            'Anticipo': anticipo,
            'SaldoRestante': saldo_restante,
            'ProductosVendidos': productos_vendidos,
            'CantidadVendida': total_cantidad,
            'PrecioCompra': precio_compra,
            'PrecioVenta': precio_venta,
            'Ganancia': total_ganancia,
        })

    return JsonResponse(data, safe=False)





@login_required(login_url="/accounts/login/")
def generar_reporte_excel_contado(request, start_date, end_date):
    start_date = timezone.datetime.strptime(start_date, '%Y-%m-%d')
    end_date = timezone.datetime.strptime(end_date, '%Y-%m-%d')

    # Filtrar ventas donde el anticipo sea igual al total de la venta
    ventas = Venta.objects.filter(
        fecha__range=[start_date, end_date],
        anticipo=F('total')
    )

    if not ventas:
    
        messages.success(
            request, '¡No hay ventas al contado En la Fecha seleccionada!', extra_tags="danger")
        return redirect('Apps.ventas:VreportC')

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=reporte_ventas_contado.xlsx'

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Reporte de Ventas"

    headers = ['Usuario', 'Fecha', 'Cliente', 'Anticipo', 'Productos Vendidos', 'Cantidad Vendida', 'Precio de Compra', 'Precio de Venta', 'Ganancia', 'Total']

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = worksheet.cell(row=2, column=col_num, value=header)

        cell.fill = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid')
        cell.border = Border(left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'),
                            top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'))
        cell.alignment = Alignment(horizontal='center')

    total_ventas = 0 
    total_ganancia_general = 0  # Variable para almacenar la suma de todos los totales

    for row_num, venta in enumerate(ventas, 3):
        detalles_venta = DetalleVenta.objects.filter(id_venta=venta)


        total_cantidad = detalles_venta.aggregate(Sum('cantidad'))['cantidad__sum']
        total_ganancia = detalles_venta.aggregate(Sum('ganancia'))['ganancia__sum']

        total_ganancia_general += total_ganancia

        user_info = f"{venta.user.username}"
        worksheet.cell(row=row_num, column=1, value=user_info).alignment = Alignment(horizontal='left', wrap_text=True)

        worksheet.cell(row=row_num, column=2, value=venta.fecha).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=3, value=venta.id_clientes.nombre).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=10, value=venta.total).alignment = Alignment(horizontal='center')

        total_ventas += venta.total  # Agregar el total de esta venta al total general

        worksheet.cell(row=row_num, column=4, value=venta.anticipo).alignment = Alignment(horizontal='center')

        productos_vendidos = ', '.join([detalle.id_producto.nombre for detalle in detalles_venta])
        worksheet.cell(row=row_num, column=5, value=productos_vendidos).alignment = Alignment(horizontal='center')

        total_cantidad = detalles_venta.aggregate(Sum('cantidad'))['cantidad__sum']
        total_ganancia = detalles_venta.aggregate(Sum('ganancia'))['ganancia__sum']

        worksheet.cell(row=row_num, column=6, value=total_cantidad).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=9, value=total_ganancia).alignment = Alignment(horizontal='center')

        # Nuevas líneas para obtener el precio de compra y precio de venta
        precio_venta = detalles_venta.aggregate(Sum('precio'))['precio__sum']
        ganancia_total = detalles_venta.aggregate(Sum('ganancia'))['ganancia__sum']
        precio_compra = precio_venta - ganancia_total

        worksheet.cell(row=row_num, column=7, value=precio_compra).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=8, value=precio_venta).alignment = Alignment(horizontal='center')

        ultima_transaccion = Transaccion.objects.filter(id_venta=venta).order_by('-fecha').first()
        saldo_restante = ultima_transaccion.saldo_venta if ultima_transaccion else 0

    # Agregar la suma total en la columna 10
    worksheet.cell(row=row_num + 1, column=10, value=total_ventas).alignment = Alignment(horizontal='center')
    worksheet.cell(row=row_num + 1, column=10).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')


    
    worksheet.cell(row=row_num + 1, column=9, value=total_ganancia_general).alignment = Alignment(horizontal='center')
    worksheet.cell(row=row_num + 1, column=9).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    for col_num in range(1, worksheet.max_column + 1):
        max_length = 0
        column = get_column_letter(col_num)
        for cell in worksheet[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width

    workbook.save(response)

    return response




@login_required(login_url="/accounts/login/")
def generar_reporte_pdf_contado(request, start_date, end_date):
    # Convertir las fechas de inicio y fin de la cadena a objetos de fecha y hora
    start_date = timezone.datetime.strptime(start_date, '%Y-%m-%d')
    end_date = timezone.datetime.strptime(end_date, '%Y-%m-%d')

    # Filtrar las ventas en el rango de fechas proporcionado y donde el anticipo sea igual al total de la venta
    ventas = Venta.objects.filter(
        fecha__range=[start_date, end_date],
        anticipo=F('total')
    )

    # Crear un objeto de respuesta HTTP con el tipo de contenido apropiado para un PDF
    response = HttpResponse(content_type='application/pdf')
    # Establecer el encabezado Content-Disposition para que el navegador abra el PDF en línea o lo descargue como un archivo adjunto
    response['Content-Disposition'] = f'attachment; filename="reporte_ventas_contado.pdf"'


    # Crear un documento PDF con el tamaño de página obtenido
    pdf = SimpleDocTemplate(response, pagesize=letter)
    # Crear una lista para almacenar los elementos del PDF
    elements = []

    # Crear la tabla de datos para las ventas
    data = []
    # Encabezados de la tabla
    data.append(['Usuario', 'Fecha', 'Cliente', 'Total', 'Anticipo', 'Productos Vendidos', 'Cantidad Vendida', 'Precio de Compra', 'Precio de Venta', 'Ganancia'])

    for venta in ventas:
        # Obtener detalles de la venta
        detalles_venta = DetalleVenta.objects.filter(id_venta=venta)

        # Obtener información del usuario
        user_info = venta.user.username

        # Obtener el nombre del cliente
        cliente_nombre = venta.id_clientes.nombre

        # Obtener los productos vendidos
        productos_vendidos = ', '.join([detalle.id_producto.nombre for detalle in detalles_venta])

        # Calcular el total de la cantidad vendida y la ganancia
        total_cantidad = detalles_venta.aggregate(Sum('cantidad'))['cantidad__sum']
        total_ganancia = detalles_venta.aggregate(Sum('ganancia'))['ganancia__sum']

        # Calcular el precio de compra y el precio de venta
        precio_venta = detalles_venta.aggregate(Sum('precio'))['precio__sum']
        ganancia_total = detalles_venta.aggregate(Sum('ganancia'))['ganancia__sum']
        precio_compra = precio_venta - ganancia_total

        # Obtener la última transacción para calcular el saldo restante
        ultima_transaccion = Transaccion.objects.filter(id_venta=venta).order_by('-fecha').first()

        # Agregar una fila de datos a la tabla
        data.append([user_info, str(venta.fecha), cliente_nombre, str(venta.total), str(venta.anticipo), productos_vendidos, str(total_cantidad), str(precio_compra), str(precio_venta), str(total_ganancia)])

    # Calcular la suma total de todos los totales
    total_ventas = sum(venta.total for venta in ventas)

    # Agregar la suma total al final de la columna 'Total' en la lista de datos
    data.append(['', '', '', f'Total de ventas: {total_ventas}', '', '', '', '', '', ''])

    # Establecer el ancho de las columnas individualmente
    column_widths = [1.5 * inch, 1 * inch, 1.5 * inch, 1 * inch, 1 * inch, 1.5 * inch, 1 * inch, 1 * inch, 1 * inch, 1 * inch]

    # Crear la tabla y agregar los datos
    tabla_data = []
    for row in data:
        tabla_row = []
        for item in row:
            # Escapar caracteres especiales para evitar errores en la generación del PDF
            item = escape(item)
            # Crear un objeto Paragraph para permitir el ajuste del texto
            tabla_row.append(Paragraph(item, getSampleStyleSheet()['Normal']))
        tabla_data.append(tabla_row)

    tabla = Table(tabla_data, colWidths=column_widths)

    # Establecer estilo para la tabla
    estilo_tabla = TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.gray),
                               ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                               ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                               ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                               ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                               ('BACKGROUND', (0, 1), (-1, -2), colors.beige),  # Color de fondo para celdas de datos
                               ('GRID', (0, 0), (-1, -1), 1, colors.black)])

    tabla.setStyle(estilo_tabla)
    elements.append(tabla)

    # Construir el PDF
    pdf.build(elements)

    return response







#---------------------------------VENTAS AL CREDITO---------------------------


@login_required(login_url="/accounts/login/")
def VentasCRView(request):

    return render(request, "ventas/VreportCR.html")


@login_required(login_url="/accounts/login/")
def obtener_datos_filtrados_credito(request, start_date, end_date):
    start_date = timezone.datetime.strptime(start_date, '%Y-%m-%d')
    end_date = timezone.datetime.strptime(end_date, '%Y-%m-%d')

    # Filtrar las ventas donde el total y el anticipo sean distintos
    ventas = Venta.objects.filter(fecha__range=[start_date, end_date]).exclude(total=F('anticipo'))

    data = []

    for venta in ventas:
        detalles_venta = DetalleVenta.objects.filter(id_venta=venta)

        user_info = f"{venta.user.username}"
        fecha = venta.fecha.strftime('%Y-%m-%d')
        cliente = venta.id_clientes.nombre
        total = venta.total
        anticipo = venta.anticipo

        productos_vendidos = ', '.join([detalle.id_producto.nombre for detalle in detalles_venta])

        total_cantidad = detalles_venta.aggregate(Sum('cantidad'))['cantidad__sum']
        total_ganancia = detalles_venta.aggregate(Sum('ganancia'))['ganancia__sum']

        precio_venta = detalles_venta.aggregate(Sum('precio'))['precio__sum']
        ganancia_total = detalles_venta.aggregate(Sum('ganancia'))['ganancia__sum']
        precio_compra = precio_venta - ganancia_total

        ultima_transaccion = Transaccion.objects.filter(id_venta=venta).order_by('-fecha').first()
        saldo_restante = ultima_transaccion.saldo_venta if ultima_transaccion else 0

        data.append({
            'Usuario': user_info,
            'Fecha': fecha,
            'Cliente': cliente,
            'Total': total,
            'Anticipo': anticipo,
            'SaldoRestante': saldo_restante,
            'ProductosVendidos': productos_vendidos,
            'CantidadVendida': total_cantidad,
            'PrecioCompra': precio_compra,
            'PrecioVenta': precio_venta,
            'Ganancia': total_ganancia,
        })

    return JsonResponse(data, safe=False)




@login_required(login_url="/accounts/login/")
def generar_reporte_excel_credito(request, start_date, end_date):
    start_date = timezone.datetime.strptime(start_date, '%Y-%m-%d')
    end_date = timezone.datetime.strptime(end_date, '%Y-%m-%d')

    # Filtrar las ventas donde el total es diferente del anticipo, es decir, ventas a crédito
    ventas = Venta.objects.filter(fecha__range=[start_date, end_date]).exclude(total=F('anticipo'))

    if not ventas:
        # Si no hay ventas al crédito en el rango de fechas especificado, puedes manejarlo como prefieras, por ejemplo, con un mensaje de error o simplemente sin retornar nada
        messages.success(request, 'No hay ventas al crédito en el rango de fechas especificado.', extra_tags="danger")
        return redirect('Apps.ventas:VreportCR')  # Redirige a la página de reporte u otra página que desees

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=reporte_ventas_credito.xlsx'

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Reporte de Ventas a Crédito"

    headers = ['Usuario', 'Fecha', 'Cliente', 'Anticipo', 'Productos Vendidos', 'Cantidad Vendida', 'Precio de Compra', 'Precio de Venta', 'Ganancia', 'Total']

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = worksheet.cell(row=2, column=col_num, value=header)

        cell.fill = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid')
        cell.border = Border(left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'),
                            top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'))
        cell.alignment = Alignment(horizontal='center')

    total_ventas = 0  # Variable para almacenar la suma de todos los totales
    total_ganancia_general = 0 

 # Mueve esta línea aquí

    

    for row_num, venta in enumerate(ventas, 3):
        detalles_venta = DetalleVenta.objects.filter(id_venta=venta)

         
        total_cantidad = detalles_venta.aggregate(Sum('cantidad'))['cantidad__sum']
        total_ganancia = detalles_venta.aggregate(Sum('ganancia'))['ganancia__sum']

        total_ganancia_general += total_ganancia 

        user_info = f"{venta.user.username}"
        worksheet.cell(row=row_num, column=1, value=user_info).alignment = Alignment(horizontal='left', wrap_text=True)

        worksheet.cell(row=row_num, column=2, value=venta.fecha).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=3, value=venta.id_clientes.nombre).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=10, value=venta.total).alignment = Alignment(horizontal='center')

        total_ventas += venta.total  # Agregar el total de esta venta al total general

        worksheet.cell(row=row_num, column=4, value=venta.anticipo).alignment = Alignment(horizontal='center')

        productos_vendidos = ', '.join([detalle.id_producto.nombre for detalle in detalles_venta])
        worksheet.cell(row=row_num, column=5, value=productos_vendidos).alignment = Alignment(horizontal='center')

        total_cantidad = detalles_venta.aggregate(Sum('cantidad'))['cantidad__sum']
        total_ganancia = detalles_venta.aggregate(Sum('ganancia'))['ganancia__sum']

        worksheet.cell(row=row_num, column=6, value=total_cantidad).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=9, value=total_ganancia).alignment = Alignment(horizontal='center')

        # Nuevas líneas para obtener el precio de compra y precio de venta
        precio_venta = detalles_venta.aggregate(Sum('precio'))['precio__sum']
        ganancia_total = detalles_venta.aggregate(Sum('ganancia'))['ganancia__sum']
        precio_compra = precio_venta - ganancia_total

        worksheet.cell(row=row_num, column=7, value=precio_compra).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=8, value=precio_venta).alignment = Alignment(horizontal='center')



        ultima_transaccion = Transaccion.objects.filter(id_venta=venta).order_by('-fecha').first()
        saldo_restante = ultima_transaccion.saldo_venta if ultima_transaccion else 0

    # Agregar la suma total en la columna 10
    worksheet.cell(row=row_num + 1, column=10, value=total_ventas).alignment = Alignment(horizontal='center')
    worksheet.cell(row=row_num + 1, column=10).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    worksheet.cell(row=row_num + 1, column=9, value=total_ganancia_general).alignment = Alignment(horizontal='center')
    worksheet.cell(row=row_num + 1, column=9).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    for col_num in range(1, worksheet.max_column + 1):
        max_length = 0
        column = get_column_letter(col_num)
        for cell in worksheet[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width

    workbook.save(response)

    return response




@login_required(login_url="/accounts/login/")
def generar_reporte_pdf_credito(request, start_date, end_date):
    # Convertir las fechas de inicio y fin de la cadena a objetos de fecha y hora
    start_date = timezone.datetime.strptime(start_date, '%Y-%m-%d')
    end_date = timezone.datetime.strptime(end_date, '%Y-%m-%d')

    # Filtrar las ventas en el rango de fechas proporcionado donde el total sea diferente al anticipo
    ventas = Venta.objects.filter(fecha__range=[start_date, end_date]).exclude(total=F('anticipo'))

    # Crear un objeto de respuesta HTTP con el tipo de contenido apropiado para un PDF
    response = HttpResponse(content_type='application/pdf')
    # Establecer el encabezado Content-Disposition para que el navegador abra el PDF en línea o lo descargue como un archivo adjunto
    response['Content-Disposition'] = f'attachment; filename="reporte_ventas_credito.pdf"'

    # Crear un documento PDF con el tamaño de página obtenido
    pdf = SimpleDocTemplate(response, pagesize=letter)
    # Crear una lista para almacenar los elementos del PDF
    elements = []

    # Crear la tabla de datos para las ventas
    data = []
    # Encabezados de la tabla
    data.append(['Usuario', 'Fecha', 'Cliente', 'Total', 'Anticipo', 'Productos Vendidos', 'Cantidad Vendida', 'Precio de Compra', 'Precio de Venta', 'Ganancia'])

    for venta in ventas:
        # Obtener detalles de la venta
        detalles_venta = DetalleVenta.objects.filter(id_venta=venta)

        # Obtener información del usuario
        user_info = venta.user.username

        # Obtener el nombre del cliente
        cliente_nombre = venta.id_clientes.nombre

        # Obtener los productos vendidos
        productos_vendidos = ', '.join([detalle.id_producto.nombre for detalle in detalles_venta])

        # Calcular el total de la cantidad vendida y la ganancia
        total_cantidad = detalles_venta.aggregate(Sum('cantidad'))['cantidad__sum']
        total_ganancia = detalles_venta.aggregate(Sum('ganancia'))['ganancia__sum']

        # Calcular el precio de compra y el precio de venta
        precio_venta = detalles_venta.aggregate(Sum('precio'))['precio__sum']
        ganancia_total = detalles_venta.aggregate(Sum('ganancia'))['ganancia__sum']
        precio_compra = precio_venta - ganancia_total

        # Agregar una fila de datos a la tabla
        data.append([user_info, str(venta.fecha), cliente_nombre, str(venta.total), str(venta.anticipo), productos_vendidos, str(total_cantidad), str(precio_compra), str(precio_venta), str(total_ganancia)])

    # Calcular la suma total de todos los totales
    total_ventas = sum(venta.total for venta in ventas)

    # Agregar la suma total al final de la columna 'Total' en la lista de datos
    data.append(['', '', '', f'Total de ventas a crédito: {total_ventas}', '', '', '', '', '', ''])

    # Establecer el ancho de las columnas individualmente
    column_widths = [1.5 * inch, 1 * inch, 1.5 * inch, 1 * inch, 1 * inch, 1.5 * inch, 1 * inch, 1 * inch, 1 * inch, 1 * inch]

    # Crear la tabla y agregar los datos
    tabla_data = []
    for row in data:
        tabla_row = []
        for item in row:
            # Escapar caracteres especiales para evitar errores en la generación del PDF
            item = escape(item)
            # Crear un objeto Paragraph para permitir el ajuste del texto
            tabla_row.append(Paragraph(item, getSampleStyleSheet()['Normal']))
        tabla_data.append(tabla_row)

    tabla = Table(tabla_data, colWidths=column_widths)

    # Establecer estilo para la tabla
    estilo_tabla = TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.gray),
                               ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                               ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                               ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                               ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                               ('BACKGROUND', (0, 1), (-1, -2), colors.beige),  # Color de fondo para celdas de datos
                               ('GRID', (0, 0), (-1, -1), 1, colors.black)])

    tabla.setStyle(estilo_tabla)
    elements.append(tabla)

    # Construir el PDF
    pdf.build(elements)

    return response






#-----------------SALDOS-------------------



@login_required(login_url="/accounts/login/")
def ClientesSView(request):

    return render(request, "ventas/VreportS.html")


@login_required(login_url="/accounts/login/")
def generar_reporte_saldos_clientes_excel(request, id_cliente):
    cliente = get_object_or_404(Cliente, pk=id_cliente)

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename=reporte_saldos_{cliente.nombre}.xlsx'

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = f"Reporte de Saldos {cliente.nombre}"

    title_font = Font(size=14, color='FFFFFF', bold=True)
    title_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    title_border = Border(left=Side(border_style='thin', color='000000'),
                          right=Side(border_style='thin', color='000000'),
                          top=Side(border_style='thin', color='000000'),
                          bottom=Side(border_style='thin', color='000000'))

    headers = ['Cliente', 'Saldo Inicial', 'Abono Saldo', 'Estado']

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = worksheet.cell(row=1, column=col_num, value=header)

        cell.fill = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid')
        cell.border = Border(left=Side(border_style='thin', color='000000'),
                             right=Side(border_style='thin', color='000000'),
                             top=Side(border_style='thin', color='000000'),
                             bottom=Side(border_style='thin', color='000000'))
        cell.alignment = Alignment(horizontal='center')

    saldos_clientes = SaldoInicial.objects.filter(cliente=cliente)
    row_num = 2

    for saldo_cliente in saldos_clientes:
        worksheet.cell(row=row_num, column=1, value=saldo_cliente.cliente.nombre).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=2, value=saldo_cliente.saldo_inicial).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=3, value=saldo_cliente.abono_saldo).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=4, value='Activo' if saldo_cliente.estado else 'Inactivo').alignment = Alignment(horizontal='center')

        row_num += 1

    for col_num in range(1, worksheet.max_column + 1):
        max_length = 0
        column = get_column_letter(col_num)
        for cell in worksheet[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width

    workbook.save(response)
    return response



@login_required(login_url="/accounts/login/")
def generar_reporte_saldos_clientes_pdf(request, id_cliente):
    cliente = get_object_or_404(Cliente, pk=id_cliente)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=reporte_saldos_{cliente.nombre}.pdf'

    pdf = SimpleDocTemplate(response, pagesize=letter)
    styles = getSampleStyleSheet()

    data = []
    headers = ['Cliente', 'Saldo Inicial', 'Abono Saldo', 'Estado']
    data.append(headers)

    saldos_clientes = SaldoInicial.objects.filter(cliente=cliente)

    for saldo_cliente in saldos_clientes:
        data.append([
            saldo_cliente.cliente.nombre,
            saldo_cliente.saldo_inicial,
            saldo_cliente.abono_saldo,
            'Activo' if saldo_cliente.estado else 'Inactivo'
        ])

    table = Table(data)
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.gray),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    table.setStyle(style)

    elements = [table]
    pdf.build(elements)
    
    return response




@login_required(login_url="/accounts/login/")
def obtener_datos_saldos_filtrados(request, id_cliente):
    cliente = get_object_or_404(Cliente, pk=id_cliente)

    datos_filtrados = []

    saldos_clientes = SaldoInicial.objects.filter(cliente=cliente)

    for saldo_cliente in saldos_clientes:
        datos_filtrados.append({
            'Cliente': saldo_cliente.cliente.nombre,
            'Saldo Inicial': saldo_cliente.saldo_inicial,
            'Abono Saldo': saldo_cliente.abono_saldo,
            'Estado': 'Activo' if saldo_cliente.estado else 'Inactivo',
        })

    return JsonResponse(datos_filtrados, safe=False)